from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

def wrap(_dir):
    workbook = load_workbook(_dir)
    worksheet = workbook.active

    dimensions = {}
    for row_index, row in enumerate(worksheet.rows):
        for column_index, cell in enumerate(row):
            if cell.value:
                dimensions[cell.column_letter] = max(
                    (dimensions.get(cell.column_letter, 0),
                    len(str(cell.value)))
                )
            
            align = 'left'
            if column_index == 0 and row_index > 2:
                align = 'right'
            
            cell.alignment = Alignment(horizontal=align)

    for column, value in dimensions.items():
        worksheet.column_dimensions[column].width = value
    
    workbook.save(_dir) 