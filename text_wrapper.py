from openpyxl import load_workbook
from openpyxl.styles import Alignment

def wrap(dir):
    workbook = load_workbook(dir)

    for rows in workbook.active.iter_rows():
        for cell in rows:
            cell.alignment = Alignment(
                horizontal='general',
                vertical='bottom',
                text_rotation=0,
                wrap_text=False,
                shrink_to_fit=False,
                indent=0
            )
    
    workbook.save(dir) 